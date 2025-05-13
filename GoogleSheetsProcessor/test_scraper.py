#!/usr/bin/env python3
# test_scraper.py - Test the scraper and output file generation

import os
import sys
import pandas as pd
import traceback
import openpyxl
from openpyxl.styles import Alignment
from datetime import datetime

# Import the debug scraper
from debug_scraper import debug_scrape_katom

def run_test():
    print("Starting scraper test...")
    
    # Test with a specific model and prefix
    model_number = "64900K"
    prefix = "150"
    
    print(f"Testing with model: {model_number}, prefix: {prefix}")
    
    try:
        # Run the debug scraper
        title, description, specs_data, specs_html, video_links, main_image, additional_images = debug_scrape_katom(model_number, prefix)
        
        if title == "Title not found" or "not found" in title.lower():
            print(f"ERROR: Could not find product. Trying a different model...")
            
            # Try a different model
            model_number = "50210"
            prefix = "731"
            print(f"Testing with model: {model_number}, prefix: {prefix}")
            
            title, description, specs_data, specs_html, video_links, main_image, additional_images = debug_scrape_katom(model_number, prefix)
            
            if title == "Title not found" or "not found" in title.lower():
                print(f"ERROR: Still could not find product. Please check the model numbers and prefixes.")
                return False
        
        # If we made it here, we found a product! Create a DataFrame and save it
        print(f"Found product: {title}")
        
        # Format the description with table at the bottom
        combined_description = f'<div style="text-align: justify;">{description}</div>'
        
        # Add the specs table below the description if it exists
        if specs_html and len(specs_html) > 0:
            combined_description += f'<h3 style="margin-top: 15px;">Specifications</h3>{specs_html}'
        
        # Create a DataFrame with the scraped data
        columns = ["Mfr Model", "Title", "Description"]
        
        # Add common spec fields
        common_spec_fields = [
            "manufacturer", "food type", "frypot style", "heat", "hertz", "nema", 
            "number of fry pots", "oil capacity/fryer (lb)", "phase", "product", 
            "product type", "rating", "special features", "type", "voltage", 
            "warranty", "weight"
        ]
        
        for field in common_spec_fields:
            columns.append(field.title())  # Title case the field names for Excel
        
        # Add video link columns
        for i in range(1, 6):  # Video Link 1, Video Link 2, etc.
            columns.append(f"Video Link {i}")
        
        # Add image columns
        columns.append("Main Image")
        for i in range(1, 6):  # Additional Image 1, Additional Image 2, etc.
            columns.append(f"Additional Image {i}")
        
        # Create row data
        row_data = {
            "Mfr Model": model_number,
            "Title": title,
            "Description": combined_description
        }
        
        # Add spec fields
        for field in common_spec_fields:
            row_data[field.title()] = specs_data.get(field, "")
        
        # Add video links
        video_list = [link.strip() for link in video_links.strip().split('\n') if link.strip()]
        for i in range(1, 6):
            if i <= len(video_list):
                row_data[f"Video Link {i}"] = video_list[i-1]
            else:
                row_data[f"Video Link {i}"] = ""
        
        # Add images
        row_data["Main Image"] = main_image
        for i in range(1, 6):
            if i <= len(additional_images):
                row_data[f"Additional Image {i}"] = additional_images[i-1]
            else:
                row_data[f"Additional Image {i}"] = ""
        
        # Create DataFrame
        df = pd.DataFrame([row_data], columns=columns)
        
        # Output path
        output_dir = os.path.expanduser("~/GoogleDriveMount/Web/Completed/Final")
        os.makedirs(output_dir, exist_ok=True)  # Ensure directory exists
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(output_dir, f"test_output_{timestamp}.xlsx")
        
        # Save to Excel
        print(f"Saving to Excel file: {output_path}")
        df.to_excel(output_path, index=False)
        
        # Adjust cell formatting with openpyxl
        print("Adjusting cell formatting...")
        workbook = openpyxl.load_workbook(output_path)
        worksheet = workbook.active
        
        # Set default row height for all rows
        for row in worksheet.iter_rows():
            worksheet.row_dimensions[row[0].row].height = 15
        
        # Adjust the wrap text settings for the Description column
        for row in worksheet.iter_rows():
            for cell in row:
                col_name = worksheet.cell(row=1, column=cell.column).value
                if col_name == "Description":
                    cell.alignment = Alignment(wrap_text=True)
        
        # Save the modified workbook
        workbook.save(output_path)
        workbook.close()
        
        print(f"Success! Output file created: {output_path}")
        print(f"Please check the file to verify that all data was scraped and saved correctly.")
        return True
        
    except Exception as e:
        print(f"Error during test: {e}")
        print(traceback.format_exc())
        return False

if __name__ == "__main__":
    run_test()
