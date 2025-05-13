#!/usr/bin/env python3
# Save this as excel_formatter.py

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
import pandas as pd
import traceback

class ExcelFormatter:
    """Helper class to enhance Excel output formatting"""
    
    def __init__(self, config_manager=None):
        self.config_manager = config_manager
        
        # Default styling
        self.header_fill = PatternFill(start_color="4285F4", end_color="4285F4", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        self.border = Border(
            left=Side(style='thin', color="CCCCCC"),
            right=Side(style='thin', color="CCCCCC"),
            top=Side(style='thin', color="CCCCCC"),
            bottom=Side(style='thin', color="CCCCCC")
        )
        
        # Column widths (in characters)
        self.column_widths = {
            "Mfr Model": 15,
            "Title": 40,
            "Description": 60,
            "Manufacturer": 20,
            "Weight": 15
        }
        
        # Default row height (in points)
        self.default_row_height = 15
        self.description_row_height = 60
    
    def format_excel_file(self, filepath):
        """Apply enhanced formatting to an Excel file"""
        try:
            if not os.path.exists(filepath):
                print(f"File not found: {filepath}")
                return False
            
            # Load the workbook
            workbook = openpyxl.load_workbook(filepath)
            worksheet = workbook.active
            
            # Format header row
            self._format_header_row(worksheet)
            
            # Format data rows
            self._format_data_rows(worksheet)
            
            # Adjust column widths
            self._adjust_column_widths(worksheet)
            
            # Save the formatted workbook
            workbook.save(filepath)
            workbook.close()
            
            print(f"Successfully formatted Excel file: {filepath}")
            return True
            
        except Exception as e:
            print(f"Error formatting Excel file: {e}")
            print(traceback.format_exc())
            return False
    
    def _format_header_row(self, worksheet):
        """Format the header row of the worksheet"""
        for cell in worksheet[1]:
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.header_alignment
            cell.border = self.border
    
    def _format_data_rows(self, worksheet):
        """Format the data rows of the worksheet"""
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), 2):
            # Set default row height
            worksheet.row_dimensions[row_idx].height = self.default_row_height
            
            for cell in row:
                # Apply border to all cells
                cell.border = self.border
                
                # Get column name
                col_idx = cell.column
                col_name = worksheet.cell(row=1, column=col_idx).value
                
                # Format based on column type
                if col_name == "Description":
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    # Set a taller row height for description rows
                    worksheet.row_dimensions[row_idx].height = self.description_row_height
                elif col_name == "Title":
                    cell.alignment = Alignment(wrap_text=True)
                elif "Weight" in str(col_name):
                    cell.alignment = Alignment(horizontal="center")
                elif "Link" in str(col_name):
                    # Make hyperlinks blue and underlined
                    if cell.value:
                        cell.font = Font(color="0000FF", underline="single")
                        cell.hyperlink = cell.value
                else:
                    cell.alignment = Alignment(vertical="center")
    
    def _adjust_column_widths(self, worksheet):
        """Adjust column widths based on content"""
        column_names = {}
        
        # Map column indices to names
        for cell in worksheet[1]:
            column_names[cell.column] = cell.value
        
        # Set column widths
        for col_idx, col_name in column_names.items():
            col_letter = get_column_letter(col_idx)
            
            if col_name in self.column_widths:
                # Use predefined width if available
                worksheet.column_dimensions[col_letter].width = self.column_widths[col_name]
            else:
                # Otherwise, use auto width with max width cap
                max_length = 0
                for cell in worksheet[col_letter]:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                
                # Cap width at 40 characters
                adjusted_width = min(max_length + 2, 40)
                worksheet.column_dimensions[col_letter].width = adjusted_width


# Extend SheetRow's save_results method without modifying the original code
def enhance_save_results(sheet_row):
    """
    Enhances the SheetRow.save_results method to use the ExcelFormatter
    without modifying the original code.
    """
    # Store the original method
    original_save_results = sheet_row.save_results
    
    # Define the enhanced method
    def enhanced_save_results():
        # Call the original method first
        original_save_results()
        
        # Then apply additional formatting if the file exists
        if hasattr(sheet_row, 'output_path') and sheet_row.output_path:
            try:
                # Get config manager if available
                config_manager = None
                if hasattr(sheet_row.parent, 'config_manager'):
                    config_manager = sheet_row.parent.config_manager
                
                # Create formatter and format the file
                formatter = ExcelFormatter(config_manager)
                formatter.format_excel_file(sheet_row.output_path)
                
                # Update status
                sheet_row.signals.update_status.emit("Enhanced formatting applied")
            except Exception as e:
                print(f"Error applying enhanced formatting: {e}")
                print(traceback.format_exc())
    
    # Replace the original method
    sheet_row.save_results = enhanced_save_results
    
    return sheet_row
