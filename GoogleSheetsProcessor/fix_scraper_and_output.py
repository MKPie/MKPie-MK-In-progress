#!/usr/bin/env python3
# fix_scraper_and_output.py - Fix scraping and output file issues

import os
import sys
import json
import traceback
import shutil
from datetime import datetime

def fix_issues():
    print("Starting to fix scraper and output file issues...")
    
    # Get the current directory
    current_dir = os.path.dirname(os.path.abspath(__file__))
    
    # 1. Ensure the output directory exists and is writable
    output_dir = os.path.expanduser("~/GoogleDriveMount/Web/Completed/Final")
    if not os.path.exists(output_dir):
        try:
            os.makedirs(output_dir, exist_ok=True)
            print(f"Created output directory: {output_dir}")
        except Exception as e:
            print(f"ERROR: Failed to create output directory: {e}")
            print("Please create this directory manually:")
            print("mkdir -p ~/GoogleDriveMount/Web/Completed/Final")
            print("chmod 755 ~/GoogleDriveMount/Web/Completed/Final")
    else:
        print(f"Output directory exists: {output_dir}")
        # Check if it's writable
        if os.access(output_dir, os.W_OK):
            print("Output directory is writable")
        else:
            print("ERROR: Output directory is not writable!")
            print("Please fix permissions:")
            print(f"chmod 755 {output_dir}")
    
    # 2. Create a test file in the output directory to verify write access
    test_file_path = os.path.join(output_dir, "test_write_access.txt")
    try:
        with open(test_file_path, 'w') as f:
            f.write("Test write access - " + datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        print(f"Successfully wrote test file: {test_file_path}")
        # Clean up test file
        os.remove(test_file_path)
        print("Test file removed")
    except Exception as e:
        print(f"ERROR: Failed to write test file: {e}")
        print("This indicates a permission issue with the output directory.")
    
    # 3. Create the image_extractor.py file
    image_extractor_path = os.path.join(current_dir, "image_extractor.py")
    with open(image_extractor_path, 'w') as f:
        f.write("""#!/usr/bin/env python3
# image_extractor.py - Add image extraction capabilities to the scraper

import os
import re
import traceback
from selenium.webdriver.common.by import By

def extract_images(driver):
    \"\"\"Extract main image and additional images from the page\"\"\"
    main_image = ""
    additional_images = []
    
    try:
        print("Looking for images on the page...")
        
        # First attempt: Look for product images using common selectors
        product_images = driver.find_elements(By.CSS_SELECTOR, 
            ".product-image img, #product-image img, #main-image img, .main-image img, [class*='product'] img, [id*='product'] img")
        
        # Second attempt: Look for gallery images
        if not product_images:
            product_images = driver.find_elements(By.CSS_SELECTOR, 
                ".gallery img, .product-gallery img, #gallery img, [class*='gallery'] img, .carousel img")
        
        # Third attempt: Look for any large images
        if not product_images:
            all_images = driver.find_elements(By.TAG_NAME, "img")
            # Filter for larger images (likely product images, not icons)
            product_images = [img for img in all_images if 
                              img.get_attribute("width") and int(img.get_attribute("width") or 0) > 200]
        
        # Final attempt: Just get all images if nothing else worked
        if not product_images:
            product_images = driver.find_elements(By.TAG_NAME, "img")
        
        print(f"Found {len(product_images)} potential product images")
        
        # If we have images, process them
        if product_images:
            # Try to identify the main image (usually the first one or the largest)
            for img in product_images:
                src = img.get_attribute("src")
                if not src:
                    continue
                    
                # Skip small images, icons, or logos
                if src.lower().endswith(('.ico', '.svg')) or 'icon' in src.lower() or 'logo' in src.lower():
                    continue
                
                # If we don't have a main image yet, set this as main
                if not main_image:
                    main_image = src
                    print(f"Selected main image: {src}")
                else:
                    # Add other images as additional
                    if src != main_image and src not in additional_images:
                        additional_images.append(src)
                        print(f"Added additional image: {src}")
                
                # Limit to 5 additional images
                if len(additional_images) >= 5:
                    break
        
        # Look for images in the page source if nothing found
        if not main_image:
            print("Searching for images in page source...")
            page_source = driver.page_source
            # Look for image URLs in the source
            img_pattern = r'https?://[^"\']+\.(jpg|jpeg|png|gif|webp)'
            matches = re.findall(img_pattern, page_source, re.IGNORECASE)
            
            for match in matches:
                url = match[0]  # The full URL
                
                # Skip small images, icons, or logos
                if 'icon' in url.lower() or 'logo' in url.lower():
                    continue
                
                if not main_image:
                    main_image = url
                    print(f"Found main image in source: {url}")
                elif url != main_image and url not in additional_images:
                    additional_images.append(url)
                    print(f"Found additional image in source: {url}")
                
                # Limit to 5 additional images
                if len(additional_images) >= 5:
                    break
    
    except Exception as e:
        print(f"Error extracting images: {e}")
        print(traceback.format_exc())
    
    return main_image, additional_images
""")
    print(f"Created image extractor: {image_extractor_path}")
    
    # 4. Create a debug version of the scraper_katom method
    debug_scraper_path = os.path.join(current_dir, "debug_scraper.py")
    with open(debug_scraper_path, 'w') as f:
        f.write("""#!/usr/bin/env python3
# debug_scraper.py - Enhanced scraper with debugging and fixes

import os
import re
import sys
import math
import time
import traceback
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from fake_useragent import UserAgent

def debug_scrape_katom(model_number, prefix, retries=2):
    \"\"\"Enhanced version of scrape_katom with retry logic, better error handling, and debugging\"\"\"
    # Clean model number
    model_number = ''.join(e for e in model_number if e.isalnum()).upper()
    if model_number.endswith("HC"):
        model_number = model_number[:-2]
    
    url = f"https://www.katom.com/{prefix}-{model_number}.html"
    print(f"DEBUG SCRAPER: Scraping URL: {url}")
    
    # Use consistent user agent instead of random
    user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36"
    
    # Empty return values
    title, description = "Title not found", "Description not found"
    specs_data = {}
    specs_html = ""
    video_links = ""
    main_image = ""
    additional_images = []
    
    # Implement retry logic
    for attempt in range(retries + 1):
        driver = None
        try:
            # Set up Selenium
            print(f"DEBUG SCRAPER: Setting up Chrome WebDriver (attempt {attempt+1}/{retries+1})...")
            options = Options()
            options.add_argument('--headless')
            options.add_argument('--no-sandbox')
            options.add_argument('--disable-dev-shm-usage')
            options.add_argument(f'user-agent={user_agent}')
            
            driver = webdriver.Chrome(options=options)
            driver.set_page_load_timeout(30)  # Set timeout to prevent hanging
            
            # Navigate to URL
            print(f"DEBUG SCRAPER: Navigating to URL: {url}")
            driver.get(url)
            
            # Check for 404
            if "404" in driver.title or "not found" in driver.title.lower():
                print(f"DEBUG SCRAPER: Product not found at {url}")
                if driver:
                    driver.quit()
                # No need to retry for 404, it's a definitive result
                return title, description, specs_data, specs_html, video_links, main_image, additional_images
            
            # Output title for debugging
            print(f"DEBUG SCRAPER: Page title: {driver.title}")
            
            # Get title
            found_title = False
            try:
                # Try multiple selectors for the title
                title_selectors = [
                    "h1.product-name.mb-0",
                    "h1.product-name",
                    "h1[class*='product-name']",
                    "h1[class*='title']",
                    ".product-title h1",
                    ".product-title",
                    "h1"
                ]
                
                # Try each selector
                for selector in title_selectors:
                    print(f"DEBUG SCRAPER: Trying title selector: {selector}")
                    title_elements = driver.find_elements(By.CSS_SELECTOR, selector)
                    if title_elements:
                        title_element = title_elements[0]
                        title = title_element.text.strip()
                        if title:
                            found_title = True
                            print(f"DEBUG SCRAPER: Found title with selector {selector}: {title}")
                            break
                
                if not found_title:
                    print("DEBUG SCRAPER: Could not find title with any selector")
            except Exception as e:
                print(f"DEBUG SCRAPER: Error getting title: {e}")
                print(traceback.format_exc())
            
            # If we found a title, get the rest of the data
            if found_title:
                # Get description
                try:
                    print("DEBUG SCRAPER: Looking for description...")
                    desc_selectors = [
                        ".tab-content",
                        ".product-description",
                        "#product-description",
                        ".description",
                        "#description"
                    ]
                    
                    for selector in desc_selectors:
                        desc_elements = driver.find_elements(By.CSS_SELECTOR, selector)
                        if desc_elements:
                            # Try to get paragraphs from the description element
                            paragraphs = desc_elements[0].find_elements(By.TAG_NAME, "p")
                            if paragraphs:
                                filtered = [
                                    f"<p>{p.text.strip()}</p>" for p in paragraphs
                                    if p.text.strip() and not p.text.lower().startswith("*free") and "video" not in p.text.lower()
                                ]
                                if filtered:
                                    description = "".join(filtered)
                                    print(f"DEBUG SCRAPER: Found description with {len(filtered)} paragraphs")
                                    break
                    
                    # If no description found, try to get the text content
                    if description == "Description not found":
                        for selector in desc_selectors:
                            desc_elements = driver.find_elements(By.CSS_SELECTOR, selector)
                            if desc_elements:
                                text = desc_elements[0].text.strip()
                                if text:
                                    description = f"<p>{text}</p>"
                                    print(f"DEBUG SCRAPER: Found description text: {text[:50]}...")
                                    break
                except Exception as e:
                    print(f"DEBUG SCRAPER: Error getting description: {e}")
                    print(traceback.format_exc())
                
                # Extract table data and HTML
                try:
                    print("DEBUG SCRAPER: Looking for specifications table...")
                    specs_data, specs_html = extract_table_data(driver)
                    print(f"DEBUG SCRAPER: Found {len(specs_data)} specification entries")
                except Exception as e:
                    print(f"DEBUG SCRAPER: Error extracting table data: {e}")
                    print(traceback.format_exc())
                
                # Extract video links
                try:
                    print("DEBUG SCRAPER: Looking for video links...")
                    video_links = extract_video_links(driver)
                    if video_links:
                        print(f"DEBUG SCRAPER: Found video links: {video_links}")
                    else:
                        print("DEBUG SCRAPER: No video links found")
                except Exception as e:
                    print(f"DEBUG SCRAPER: Error extracting video links: {e}")
                    print(traceback.format_exc())
                
                # Extract images
                try:
                    print("DEBUG SCRAPER: Looking for images...")
                    from image_extractor import extract_images
                    main_image, additional_images = extract_images(driver)
                    if main_image:
                        print(f"DEBUG SCRAPER: Found main image: {main_image}")
                    else:
                        print("DEBUG SCRAPER: No main image found")
                        
                    if additional_images:
                        print(f"DEBUG SCRAPER: Found {len(additional_images)} additional images")
                    else:
                        print("DEBUG SCRAPER: No additional images found")
                except Exception as e:
                    print(f"DEBUG SCRAPER: Error extracting images: {e}")
                    print(traceback.format_exc())
                
                # Success! No need for more retries
                print(f"DEBUG SCRAPER: Successfully scraped {url}")
                break
                
            else:
                # Title not found, maybe retry
                if attempt < retries:
                    retry_wait = (attempt + 1) * 2  # Progressive backoff
                    print(f"DEBUG SCRAPER: Title not found. Retry {attempt+1}/{retries} in {retry_wait} seconds...")
                    time.sleep(retry_wait)
                else:
                    print(f"DEBUG SCRAPER: All retries failed for {url}")
            
        except Exception as e:
            print(f"DEBUG SCRAPER: Error in scrape attempt {attempt+1}: {e}")
            print(traceback.format_exc())
            
            # Only retry if this wasn't the last attempt
            if attempt < retries:
                retry_wait = (attempt + 1) * 2  # Progressive backoff
                print(f"DEBUG SCRAPER: Retry {attempt+1}/{retries} in {retry_wait} seconds...")
                time.sleep(retry_wait)
        
        finally:
            # Ensure driver is always closed, even if an exception occurs
            if driver:
                try:
                    driver.quit()
                    print("DEBUG SCRAPER: WebDriver closed")
                except:
                    print("DEBUG SCRAPER: Error closing WebDriver")
    
    # Print summary of what we found
    print("\nDEBUG SCRAPER RESULTS SUMMARY:")
    print(f"Title: {title}")
    print(f"Description: {description[:100]}...")
    print(f"Specs data entries: {len(specs_data)}")
    print(f"Specs HTML length: {len(specs_html)}")
    print(f"Video links: {video_links or 'None'}")
    print(f"Main image: {main_image or 'None'}")
    print(f"Additional images: {len(additional_images)}")
    
    return title, description, specs_data, specs_html, video_links, main_image, additional_images

def extract_table_data(driver):
    \"\"\"
    Extract table data both as a dictionary of key-value pairs AND as an HTML table.
    Returns a tuple: (specs_dict, specs_html)
    \"\"\"
    specs_dict = {}
    specs_html = ""
    
    try:
        # Try multiple approaches to find the table
        
        # First, try to get the original table HTML
        specs_tables = driver.find_elements(By.CSS_SELECTOR, "table.table.table-condensed.specs-table")
        
        if not specs_tables:
            # Try generic tables
            specs_tables = driver.find_elements(By.TAG_NAME, "table")
        
        if specs_tables:
            # Extract key-value pairs from the table
            table = specs_tables[0]
            rows = table.find_elements(By.TAG_NAME, "tr")
            
            # Build a clean table with slim styling
            specs_html = '<table class="specs-table" cellspacing="0" cellpadding="4" border="1" style="margin-top:10px;border-collapse:collapse;width:auto;" align="left"><tbody>'
            
            for row in rows:
                cells = row.find_elements(By.TAG_NAME, "td")
                if len(cells) >= 2:
                    key = cells[0].text.strip()
                    value = cells[1].text.strip()
                    
                    # Check if this is a weight field and process accordingly
                    if "weight" in key.lower():
                        value = process_weight_value(value)
                    
                    # Add to the dictionary
                    if key and not key.lower() in specs_dict:
                        specs_dict[key.lower()] = value
                    
                    # Add to the HTML table
                    specs_html += f'<tr><td style="padding:3px 8px;"><b>{key}</b></td><td style="padding:3px 8px;">{value}</td></tr>'
            
            specs_html += "</tbody></table>"
        
        # If no table found or no HTML extracted, create an HTML table from the data we find
        if not specs_html or specs_html == "":
            # Start building an HTML table
            other_specs = []
            
            # Try to find spec elements in various ways
            
            # Method 1: Look for dedicated spec elements
            spec_rows = driver.find_elements(By.CSS_SELECTOR, ".specs-row, [class*='spec']")
            if spec_rows:
                for row in spec_rows:
                    key_elem = row.find_elements(By.CSS_SELECTOR, ".spec-key, .spec-name, [class*='key'], [class*='name']")
                    val_elem = row.find_elements(By.CSS_SELECTOR, ".spec-value, .spec-val, [class*='value'], [class*='val']")
                    
                    if key_elem and val_elem:
                        key = key_elem[0].text.strip()
                        value = val_elem[0].text.strip()
                        
                        # Check if this is a weight field and process accordingly
                        if "weight" in key.lower():
                            value = process_weight_value(value)
                            
                        if key:
                            other_specs.append((key, value))
                            if not key.lower() in specs_dict:
                                specs_dict[key.lower()] = value
            
            # Method 2: Look for definition lists
            if not other_specs:
                dl_elements = driver.find_elements(By.TAG_NAME, "dl")
                for dl in dl_elements:
                    terms = dl.find_elements(By.TAG_NAME, "dt")
                    definitions = dl.find_elements(By.TAG_NAME, "dd")
                    
                    for i in range(min(len(terms), len(definitions))):
                        key = terms[i].text.strip()
                        value = definitions[i].text.strip()
                        
                        # Check if this is a weight field and process accordingly
                        if "weight" in key.lower():
                            value = process_weight_value(value)
                            
                        if key:
                            other_specs.append((key, value))
                            if not key.lower() in specs_dict:
                                specs_dict[key.lower()] = value
            
            # Method 3: Look for text patterns in all content
            if not other_specs:
                # Get all elements that might contain specs
                elements = driver.find_elements(By.CSS_SELECTOR, "p, div, li, span")
                
                # Common spec terms to look for - expand this list as needed
                common_specs = [
                    "manufacturer", "food type", "frypot style", "heat", "hertz", "nema", 
                    "number of", "oil capacity", "phase", "product", "type", "rating", 
                    "special features", "voltage", "warranty", "weight", "dimensions"
                ]
                
                for element in elements:
                    text = element.text.strip()
                    if not text or len(text) > 100:  # Skip empty or very long text
                        continue
                    
                    # Look for patterns like "Key: Value" or "Key - Value"
                    for pattern in [r'([^:]+):\s*(.+)', r'([^-]+)-\s*(.+)']: 
                        match = re.match(pattern, text)
                        if match:
                            key = match.group(1).strip()
                            value = match.group(2).strip()
                            
                            # Check if this is a weight field and process accordingly
                            if "weight" in key.lower():
                                value = process_weight_value(value)
                            
                            # Check if this key is one of our common specs
                            if any(spec in key.lower() for spec in common_specs):
                                other_specs.append((key, value))
                                if not key.lower() in specs_dict:
                                    specs_dict[key.lower()] = value
                                break
            
            # Create HTML table from the data we collected
            if other_specs:
                specs_html = '<table class="specs-table" cellspacing="0" cellpadding="4" border="1" style="margin-top:10px;border-collapse:collapse;width:auto;" align="left"><tbody>'
                for key, value in other_specs:
                    specs_html += f'<tr><td style="padding:3px 8px;"><b>{key}</b></td><td style="padding:3px 8px;">{value}</td></tr>'
                specs_html += "</tbody></table>"
    
    except Exception as e:
        print(f"Error extracting table data: {e}")
        print(traceback.format_exc())
    
    return specs_dict, specs_html

def extract_video_links(driver):
    \"\"\"Extract video links from the page\"\"\"
    video_links = ""
    
    try:
        # Find source tags with .mp4 files
        sources = driver.find_elements(By.CSS_SELECTOR, "source[src*='.mp4'], source[type*='video']")
        for source in sources:
            src = source.get_attribute("src")
            if src and src not in video_links:
                video_links += f"{src}\\n"
        
        # If no video sources found, look for video elements
        if not video_links:
            videos = driver.find_elements(By.TAG_NAME, "video")
            for video in videos:
                # Try to get source elements within video tag
                inner_sources = video.find_elements(By.TAG_NAME, "source")
                for source in inner_sources:
                    src = source.get_attribute("src")
                    if src and src not in video_links:
                        video_links += f"{src}\\n"
                        
        # Last resort - extract video URLs from the page source
        if not video_links:
            page_source = driver.page_source
            # Look for .mp4 URLs in the source
            mp4_pattern = r'https?://[^"\']+\.mp4'
            matches = re.findall(mp4_pattern, page_source)
            for match in matches:
                if match not in video_links:
                    video_links += f"{match}\\n"
    except Exception as e:
        print(f"Error extracting video links: {e}")
        print(traceback.format_exc())
    
    return video_links

def process_weight_value(value):
    \"\"\"Process weight values: round up to whole number and add 5\"\"\"
    try:
        # Try to extract a number from the string
        # This handles cases like "22.93" or "22.93 lbs"
        number_match = re.search(r'(\\d+(\\.\\d+)?)', str(value))
        if number_match:
            # Extract the number
            number = float(number_match.group(1))
            
            # Round up to nearest whole number
            rounded = math.ceil(number)
            
            # Add 5
            final = rounded + 5
            
            # If the original had units, keep them
            units_match = re.search(r'[^\\d.]+$', str(value))
            units = units_match.group(0).strip() if units_match else ""
            
            return f"{final}{' ' + units if units else ''}"
        return value
    except:
        # If any error occurs, return the original value
        return value

if __name__ == "__main__":
    print("This script is meant to be imported into main.py")
    print("Example usage:")
    print("from debug_scraper import debug_scrape_katom")
    print("title, description, specs_data, specs_html, video_links, main_image, additional_images = debug_scrape_katom(model_number, prefix)")
""")
    print(f"Created debug scraper: {debug_scraper_path}")
    
    # 5. Create a test script that will scrape a specific product and output the results
    test_script_path = os.path.join(current_dir, "test_scraper.py")
    with open(test_script_path, 'w') as f:
        f.write("""#!/usr/bin/env python3
# test_scraper.py - Test the scraper and output file generation

import os
import sys
import pandas as pd
import traceback
import openpyxl
from openpyxl.styles import Alignment
from datetime import datetime

# Import the debug scraper
from debug_scraper import debug_scrape_katom

def run_test():
    print("Starting scraper test...")
    
    # Test with a specific model and prefix
    model_number = "64900K"
    prefix = "150"
    
    print(f"Testing with model: {model_number}, prefix: {prefix}")
    
    try:
        # Run the debug scraper
        title, description, specs_data, specs_html, video_links, main_image, additional_images = debug_scrape_katom(model_number, prefix)
        
        if title == "Title not found" or "not found" in title.lower():
            print(f"ERROR: Could not find product. Trying a different model...")
            
            # Try a different model
            model_number = "50210"
            prefix = "731"
            print(f"Testing with model: {model_number}, prefix: {prefix}")
            
            title, description, specs_data, specs_html, video_links, main_image, additional_images = debug_scrape_katom(model_number, prefix)
            
            if title == "Title not found" or "not found" in title.lower():
                print(f"ERROR: Still could not find product. Please check the model numbers and prefixes.")
                return False
        
        # If we made it here, we found a product! Create a DataFrame and save it
        print(f"Found product: {title}")
        
        # Format the description with table at the bottom
        combined_description = f'<div style="text-align: justify;">{description}</div>'
        
        # Add the specs table below the description if it exists
        if specs_html and len(specs_html) > 0:
            combined_description += f'<h3 style="margin-top: 15px;">Specifications</h3>{specs_html}'
        
        # Create a DataFrame with the scraped data
        columns = ["Mfr Model", "Title", "Description"]
        
        # Add common spec fields
        common_spec_fields = [
            "manufacturer", "food type", "frypot style", "heat", "hertz", "nema", 
            "number of fry pots", "oil capacity/fryer (lb)", "phase", "product", 
            "product type", "rating", "special features", "type", "voltage", 
            "warranty", "weight"
        ]
        
        for field in common_spec_fields:
            columns.append(field.title())  # Title case the field names for Excel
        
        # Add video link columns
        for i in range(1, 6):  # Video Link 1, Video Link 2, etc.
            columns.append(f"Video Link {i}")
        
        # Add image columns
        columns.append("Main Image")
        for i in range(1, 6):  # Additional Image 1, Additional Image 2, etc.
            columns.append(f"Additional Image {i}")
        
        # Create row data
        row_data = {
            "Mfr Model": model_number,
            "Title": title,
            "Description": combined_description
        }
        
        # Add spec fields
        for field in common_spec_fields:
            row_data[field.title()] = specs_data.get(field, "")
        
        # Add video links
        video_list = [link.strip() for link in video_links.strip().split('\\n') if link.strip()]
        for i in range(1, 6):
            if i <= len(video_list):
                row_data[f"Video Link {i}"] = video_list[i-1]
            else:
                row_data[f"Video Link {i}"] = ""
        
        # Add images
        row_data["Main Image"] = main_image
        for i in range(1, 6):
            if i <= len(additional_images):
                row_data[f"Additional Image {i}"] = additional_images[i-1]
            else:
                row_data[f"Additional Image {i}"] = ""
        
        # Create DataFrame
        df = pd.DataFrame([row_data], columns=columns)
        
        # Output path
        output_dir = os.path.expanduser("~/GoogleDriveMount/Web/Completed/Final")
        os.makedirs(output_dir, exist_ok=True)  # Ensure directory exists
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(output_dir, f"test_output_{timestamp}.xlsx")
        
        # Save to Excel
        print(f"Saving to Excel file: {output_path}")
        df.to_excel(output_path, index=False)
        
        # Adjust cell formatting with openpyxl
        print("Adjusting cell formatting...")
        workbook = openpyxl.load_workbook(output_path)
        worksheet = workbook.active
        
        # Set default row height for all rows
        for row in worksheet.iter_rows():
            worksheet.row_dimensions[row[0].row].height = 15
        
        # Adjust the wrap text settings for the Description column
        for row in worksheet.iter_rows():
            for cell in row:
                col_name = worksheet.cell(row=1, column=cell.column).value
                if col_name == "Description":
                    cell.alignment = Alignment(wrap_text=True)
        
        # Save the modified workbook
        workbook.save(output_path)
        workbook.close()
        
        print(f"Success! Output file created: {output_path}")
        print(f"Please check the file to verify that all data was scraped and saved correctly.")
        return True
        
    except Exception as e:
        print(f"Error during test: {e}")
        print(traceback.format_exc())
        return False

if __name__ == "__main__":
    run_test()
""")
    print(f"Created test script: {test_script_path}")
    
    # 6. Make the scripts executable
    try:
        os.chmod(debug_scraper_path, 0o755)
        os.chmod(image_extractor_path, 0o755)
        os.chmod(test_script_path, 0o755)
        print("Made scripts executable")
    except:
        print("Could not make scripts executable - please run them with python command")
    
    print("\nAll fixes have been applied!")
    print("\nNEXT STEPS:")
    print("1. Run the test script to verify scraping and output file generation:")
    print(f"   python {test_script_path}")
    print("2. Check the output file to verify all data was scraped and saved correctly")
    print("3. If the test is successful, update your main.py to use the improved scraper")
    print("   (The debug_scraper.py contains all the improvements needed)")

if __name__ == "__main__":
    fix_issues()
